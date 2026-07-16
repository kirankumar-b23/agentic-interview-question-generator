#!/usr/bin/env python3
"""Ingest REAL GenAI interview questions from the raw NxtMock / Interview-Intelligence
spreadsheets in data/raw/ into the curated GenAI bank (data/genai_question_bank.json).

Unlike scripts/build_genai_bank.py (which harvests from the web via Tavily and REBUILDS the
bank), this reads the local raw files, keeps GenAI-only questions that carry a real company
and (where available) a verified EASY/MEDIUM/HARD difficulty, cleans/normalizes the company,
role-tags, de-duplicates, and APPENDS the new ones to the existing bank. It is idempotent —
re-running adds nothing already present (safe to run again after a Tavily re-harvest).

Sources (all under data/raw/):
  - Interview Intelligence Master_ 2026 - Master Sheet.csv     (primary; ~8.5k Qs)
  - Interview Intelligence Master_ 2026.xlsx                   (same + extra sheets)
  - Interview Intelligence Master_ Jan-Dec 2025.xlsx           (a whole extra year)
  - Fulltime_Jobs_Curriculum_v2_COMPLETE - Rejection Selection Analysis.xlsx (AI/ML rows)
  - GEN_AI_NXTMOCK_INTERVIEW.xlsx                              (product export; GEN_AI rows)

Needs: openpyxl (pandas is NOT required). No network.
Run:  python scripts/ingest_xlsx_questions.py
"""
from __future__ import annotations
import csv
import importlib.util
import json
import re
import sys
import uuid
from collections import Counter
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl  # noqa: E402

from src.config import DATA_DIR, GENAI_BANK_JSON  # noqa: E402

# Reuse helpers from the Tavily build script without requiring scripts/ to be a package.
_spec = importlib.util.spec_from_file_location(
    "build_genai_bank", Path(__file__).resolve().parent / "build_genai_bank.py")
_bgb = importlib.util.module_from_spec(_spec)
_spec.loader.exec_module(_bgb)
_infer_role = _bgb._infer_role
_norm = _bgb._norm
_CANON = _bgb._CANON
DEDUP_SIM = _bgb.DEDUP_SIM

RAW = DATA_DIR / "raw"
CSV_FILE = RAW / "Interview Intelligence Master_ 2026 - Master Sheet.csv"
XLSX_MASTERS = [
    RAW / "Interview Intelligence Master_ 2026.xlsx",
    RAW / "Interview Intelligence Master_ Jan-Dec 2025.xlsx",
]
CURRICULUM_FILE = RAW / "Fulltime_Jobs_Curriculum_v2_COMPLETE - Rejection Selection Analysis.xlsx"
NXTMOCK_FILE = RAW / "GEN_AI_NXTMOCK_INTERVIEW.xlsx"

# ── GenAI filter ──────────────────────────────────────────────────────────────────────────
# Long/unambiguous phrases match as substrings; short/ambiguous tokens match on WORD BOUNDARIES
# (so "rag" no longer matches inside "sto-rag-e"/"ave-rag-e", "ai" not inside "email", etc.).
_PHRASES = [
    "large language model", "prompt engineer", "prompt injection", "prompt template",
    "system prompt", "generative ai", "generative artificial", "generative model",
    "retrieval augmented", "retrieval-augmented", "vector database", "vector db", "vector store",
    "word embedding", "text embedding", "embedding model", "langchain", "langgraph", "llamaindex",
    "chain of thought", "chain-of-thought", "hallucinat", "fine-tun", "fine tun", "transformer",
    "attention mechanism", "self-attention", "hugging face", "diffusion", "openai", "gemini",
    "anthropic", "agentic", "ai agent", "multi-agent", "function calling", "tokeniz",
    "context window", "semantic search", "foundation model", "text-to-image", "text to speech",
    "responsible ai", "guardrail",
]
_WORD = [r"llms?", r"gpt", r"rag", r"genai", r"gen ai", r"mcp", r"nlp",
         r"few[- ]shot", r"zero[- ]shot", r"lora", r"peft", r"bert"]
_WL = [re.compile(r"(?<![a-z])(?:%s)(?![a-z])" % p) for p in _WORD]

# Aptitude / verbal / non-technical topics — NEVER GenAI, even if a stray keyword appears in the
# question text (e.g. a vocabulary MCQ whose answer option happens to be "Hallucination").
_EXCLUDE_TOPICS = {
    "vocabulary", "data_interpretation", "verbal", "aptitude", "logical_reasoning",
    "quantitative", "reasoning", "grammar", "reading_comprehension", "synonyms",
    "antonyms", "sentence_correction", "sentence_completion", "quantitative_aptitude",
}


def _hit(text: str) -> bool:
    t = (text or "").lower()
    return any(p in t for p in _PHRASES) or any(rx.search(t) for rx in _WL)


# Verbal/quantitative-aptitude question STEMS. A question matching one of these is never a GenAI
# interview question, even if a keyword like "hallucination" appears only as an MCQ answer option
# (this catches the case where the same question also exists in a master sheet with no topic).
_APTITUDE_RX = re.compile(
    r"(choose the (option|word|correct|most)|matches in meaning|opposite in meaning|"
    r"most nearly (the same|opposite)|synonym|antonym|odd one out|fill in the blank|"
    r"correctly spel|arrange the .*(words|sentences|letters)|"
    r"find the (missing|next|odd) (number|term|letter|figure))", re.IGNORECASE)


def _is_genai(question: str, topic: str = "", sub_topic: str = "") -> bool:
    """GenAI iff the signal is in the QUESTION text or a topic/sub_topic — with an absolute veto for
    aptitude/verbal topics (authoritative) and for aptitude question stems (robust across sources)."""
    if _norm(topic) in _EXCLUDE_TOPICS or _norm(sub_topic) in _EXCLUDE_TOPICS:
        return False
    if _APTITUDE_RX.search(question or ""):
        return False
    if _hit(question):
        return True
    return _hit(topic) or _hit(sub_topic)


# ── company cleaning (company-only policy stays; these are curated internal names) ──────────
# Trailing "noise" tokens peeled from the END of a company name (leaves mid-name words intact, so
# "Delta Technology and Management Services" → "Delta Technology and Management", not "Delta and
# Management"). Peeling stops as soon as a non-noise word is hit, and never empties the name.
_TRAIL = {
    "pvt", "pvt.", "private", "ltd", "ltd.", "limited", "llp", "inc", "inc.", "incorporated",
    "corp", "corp.", "corporation", "co", "co.", "company", "technologies", "technology",
    "solutions", "solution", "systems", "system", "labs", "lab", "software", "global",
    "india", "services", "service", "tech", "consulting", "enterprises",
}
_TYPO = {
    "palensto": "Palnesto",
    "trytvara": "Tvara",
    "bluepond.ai": "BluePond.AI",
    "cobuildx.ai": "CobuildX.ai",
}


def _clean_company(raw: str | None) -> str | None:
    if not raw:
        return None
    # drop anything after " - " (e.g. "CXfirst.ai - Full Stack") and stray quotes
    c = re.split(r"\s+-\s+", raw.strip())[0].strip().strip('"').strip()
    # strip annotation noise: parenthetical suffix "(AIML role)", trailing "interview process",
    # and a dangling connector/name fragment like "Yashik Yadav &"
    c = re.sub(r"\s*\([^)]*\)\s*$", "", c)
    c = re.sub(r"\s+(interview|hiring|recruitment)\s+process\b.*$", "", c, flags=re.IGNORECASE)
    c = re.sub(r"\s+interview\b.*$", "", c, flags=re.IGNORECASE)
    c = c.strip().strip("&,-. ").strip()
    if not c:
        return None
    key = c.lower()
    if key in _TYPO:
        return _TYPO[key]
    if key in _CANON:
        return _CANON[key]
    # peel trailing noise tokens, always keeping at least the first word
    tokens = [t for t in re.split(r"[\s,]+", c) if t]
    while len(tokens) > 1 and re.sub(r"[^\w.]", "", tokens[-1]).lower() in _TRAIL:
        tokens.pop()
    # drop a now-dangling trailing connector token (e.g. "Yashik Yadav & Co" → "… &" → "Yashik Yadav")
    while len(tokens) > 1 and re.sub(r"[^\w.]", "", tokens[-1]) == "":
        tokens.pop()
    result = " ".join(tokens).strip(" ,.-&")
    rkey = result.lower()
    result = _TYPO.get(rkey) or _CANON.get(rkey) or result
    # reject obvious non-companies
    if len(result) < 2 or result.lower() in {"n/a", "na", "none", "unknown", "-"}:
        return None
    return result


def _difficulty(raw: str | None) -> str:
    d = (raw or "").strip().lower()
    return {
        "easy": "Easy", "beginner": "Easy",
        "medium": "Medium", "intermediate": "Medium", "moderate": "Medium",
        "hard": "Hard", "advanced": "Hard", "expert": "Hard", "difficult": "Hard",
    }.get(d, "Medium")


# ── raw extractors → dicts {question, company, difficulty, topic, sub_topic, companies?} ────
def _hdr_index(header) -> dict:
    idx = {}
    for i, h in enumerate(header):
        if h is not None and str(h).strip():
            idx[str(h).replace("\n", " ").strip().lower()] = i
    return idx


def _cell(row, idx, name) -> str:
    i = idx.get(name)
    if i is None or i >= len(row) or row[i] is None:
        return ""
    return str(row[i]).strip()


def _rows_from_csv(path: Path):
    if not path.exists():
        return
    with open(path, newline="", encoding="utf-8", errors="ignore") as fh:
        reader = csv.reader(fh)
        header = next(reader, None)
        if not header:
            return
        idx = _hdr_index(header)
        if "question" not in idx:
            return
        for row in reader:
            q = _cell(row, idx, "question")
            if q:
                yield {"question": q, "company": _cell(row, idx, "company name"),
                       "difficulty": _cell(row, idx, "difficulty_level"),
                       "topic": _cell(row, idx, "topic"), "sub_topic": _cell(row, idx, "sub_topic")}


def _rows_from_xlsx_masters(path: Path):
    if not path.exists():
        return
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        for ws in wb.worksheets:
            it = ws.iter_rows(values_only=True)
            header = next(it, None)
            if not header:
                continue
            idx = _hdr_index(header)
            if "question" not in idx:
                continue
            for row in it:
                q = _cell(row, idx, "question")
                if q:
                    yield {"question": q, "company": _cell(row, idx, "company name"),
                           "difficulty": _cell(row, idx, "difficulty_level"),
                           "topic": _cell(row, idx, "topic"),
                           "sub_topic": _cell(row, idx, "sub_topic")}
    finally:
        wb.close()


def _split_questions(cell: str) -> list[str]:
    """Split a bundled 'Questions Asked' cell into individual questions (keep only ?-ended ones)."""
    flat = re.sub(r"\s+", " ", (cell or "").replace("\n", " ")).strip()
    parts = re.split(r"(?<=\?)\s+", flat)
    return [p.strip() for p in parts if p.strip() and p.strip().endswith("?")]


def _rows_from_curriculum(path: Path):
    if not path.exists():
        return
    wb = openpyxl.load_workbook(path, data_only=True)
    try:
        ws = wb["Curriculum Breakdown"]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return
        header = [str(c).strip() if c else "" for c in rows[0]]
        idx = {h.lower(): i for i, h in enumerate(header) if h}
        si = idx.get("subject")
        ti = idx.get("topic")
        qi = next((i for i, h in enumerate(header) if "questions asked" in h.lower()), None)
        ci = idx.get("companies")
        if None in (si, ti, qi, ci):
            return
        for r in rows[1:]:
            subj = str(r[si]).strip() if si < len(r) and r[si] else ""
            if "ai/ml" not in subj.lower():
                continue
            qcell = str(r[qi]) if qi < len(r) and r[qi] else ""
            companies = ([c.strip() for c in str(r[ci]).split(",") if c.strip()]
                         if ci < len(r) and r[ci] else [])
            topic = str(r[ti]) if ti < len(r) and r[ti] else ""
            for q in _split_questions(qcell):
                yield {"question": q, "company": companies[0] if companies else "",
                       "difficulty": "", "topic": topic, "sub_topic": "",
                       "companies": companies}
    finally:
        wb.close()


def _rows_from_nxtmock(path: Path):
    if not path.exists():
        return
    wb = openpyxl.load_workbook(path, data_only=True)
    try:
        ws = wb["QuestionDetails"]
        it = ws.iter_rows(values_only=True)
        header = next(it, None)
        if not header:
            return
        idx = _hdr_index(header)
        for row in it:
            if _cell(row, idx, "category").upper() != "GEN_AI":
                continue
            q = _cell(row, idx, "content")
            if q:
                yield {"question": q, "company": _cell(row, idx, "asked_in_company"),
                       "difficulty": _cell(row, idx, "difficulty"),
                       "topic": _cell(row, idx, "topic"), "sub_topic": _cell(row, idx, "sub_topic")}
    finally:
        wb.close()


# ── build candidate records ─────────────────────────────────────────────────────────────────
def _candidates() -> list[dict]:
    sources = [
        ("nxtmock", _rows_from_csv(CSV_FILE)),
        ("nxtmock", (r for p in XLSX_MASTERS for r in _rows_from_xlsx_masters(p))),
        ("curriculum", _rows_from_curriculum(CURRICULUM_FILE)),
        ("nxtmock", _rows_from_nxtmock(NXTMOCK_FILE)),
    ]
    out = []
    for source, rows in sources:
        for row in rows:
            q = (row.get("question") or "").strip()
            if not q or not _is_genai(q, row.get("topic", ""), row.get("sub_topic", "")):
                continue
            # company set: curriculum rows carry the whole list; others carry one
            raw_companies = row.get("companies") or ([row["company"]] if row.get("company") else [])
            companies = {c for c in (_clean_company(x) for x in raw_companies) if c}
            primary = _clean_company(row.get("company"))
            if not primary:
                primary = next(iter(companies), None)
            if not primary:                       # company-only policy: drop no-company rows
                continue
            companies.add(primary)
            out.append({
                "text": q,
                "company": primary,
                "companies": companies,
                "difficulty": _difficulty(row.get("difficulty")),
                "source": source,
            })
    return out


def _merge(a: dict, b: dict) -> None:
    """Fold b into a: union companies, keep a real (non-default) difficulty, prefer a real company."""
    a["companies"] |= b["companies"]
    if a["difficulty"] == "Medium" and b["difficulty"] != "Medium":
        a["difficulty"] = b["difficulty"]


def _dedup(cands: list[dict]) -> list[dict]:
    # 1) exact-normalized grouping
    groups: dict[str, dict] = {}
    for c in cands:
        k = _norm(c["text"])
        if not k:
            continue
        if k in groups:
            _merge(groups[k], c)
        else:
            groups[k] = c
    items = list(groups.values())

    # 2) semantic near-duplicate collapse (embeddings, else the exact grouping above stands)
    try:
        from src import embeddings
        sim = embeddings.cosine_matrix([it["text"] for it in items])
        if sim is not None:
            drop = set()
            for i in range(len(items)):
                if i in drop:
                    continue
                for j in range(i + 1, len(items)):
                    if j not in drop and sim[i][j] >= DEDUP_SIM:
                        _merge(items[i], items[j])
                        drop.add(j)
            items = [it for k, it in enumerate(items) if k not in drop]
    except Exception as e:  # noqa: BLE001
        print(f"[dedup] embeddings unavailable ({e}); used exact-text dedup only")
    return items


def main() -> int:
    if not GENAI_BANK_JSON.exists():
        print(f"ERROR: {GENAI_BANK_JSON} not found — run build_genai_bank.py first.")
        return 1
    bank = json.loads(GENAI_BANK_JSON.read_text(encoding="utf-8"))
    existing = {_norm(q["content"]) for q in bank}

    cands = _candidates()
    print(f"extracted {len(cands)} GenAI candidate rows (company-only) from raw files")
    merged = _dedup(cands)
    print(f"→ {len(merged)} distinct after dedup")

    added = []
    for it in merged:
        if _norm(it["text"]) in existing:
            continue
        existing.add(_norm(it["text"]))
        added.append({
            "id": str(uuid.uuid4()),
            "content": it["text"],
            "topic": "Gen AI",
            "difficulty": it["difficulty"],
            "company": it["company"],
            "role": _infer_role(it["text"]),
            "source": it["source"],
            "source_url": "",
            "source_count": len(it["companies"]),
        })

    bank.extend(added)
    GENAI_BANK_JSON.write_text(json.dumps(bank, indent=2, ensure_ascii=False), encoding="utf-8")

    print(f"\nAdded {len(added)} new GenAI questions → bank now {len(bank)} total")
    if added:
        print(f"  roles:      {dict(Counter(q['role'] for q in added))}")
        print(f"  difficulty: {dict(Counter(q['difficulty'] for q in added))}")
        print(f"  sources:    {dict(Counter(q['source'] for q in added))}")
        print(f"  distinct new companies: {len({q['company'] for q in added})}")
        print(f"  top companies: {Counter(q['company'] for q in added).most_common(12)}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
